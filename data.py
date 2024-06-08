# %%
import pandas as pd
import openpyxl
from openpyxl.chart import PieChart, Reference


# %%
data = pd.read_csv('drug_consumption.csv')

# %%
## Age recoding
age_col = {
          -0.95197: '18-24',
          -0.07854: '25 - 34',
          0.49788: '35 - 44',
          1.09449: '45 - 54',
          1.82213: '55 - 64',
          2.59171: '65+'
          }
data['Age'] = data['Age'].replace(age_col)

# %%
## Gender recoding
gender_col = {
            0.48246: 'Female',
            -0.48246: 'Male'
            }
data['Gender'] = data['Gender'].replace(gender_col)

# %%
## Education recoding
education_col = {
            -2.43591: 'Left School Before 16 years',
            -1.73790: 'Left School at 16 years',
            -1.43719: 'Left School at 17 years',
            -1.22751: 'Left School at 18 years',
            -0.61113: 'Some College,No Certificate Or Degree',
            -0.05921: 'Professional Certificate/ Diploma',
            0.45468: 'University Degree',
            1.16365: 'Masters Degree',
            1.98437: 'Doctorate Degree',
            }
data['Education'] = data['Education'].replace(education_col)

# %%
## Country recoding
country_col = {
            -0.09765: 'Australia',
            0.24923: 'Canada',
            -0.46841: 'New Zealand',
            -0.28519: 'Other',
            0.21128: 'Republic of Ireland',
            0.96082: 'UK',
            -0.57009: 'USA'
            }
data['Country'] = data['Country'].replace(country_col)

# %%
## Ethnicity recoding
ethnicity_col = {
            -0.50212: 'Asian',
            -1.10702: 'Black',
            1.90725: 'Mixed-Black/Asian',
            0.12600: 'Mixed-White/Asian',
            -0.22166: 'Mixed-White/Black',
            0.11440: 'Other',
            -0.31685: 'White'
            }
data['Ethnicity'] = data['Ethnicity'].replace(ethnicity_col)

# %%
## Usage recoding
usage_col = {
    'CL0': 'Never Used',
    'CL1': 'Used over a Decade Ago',
    'CL2': 'Used in Last Decade',
    'CL3': 'Used in Last Year',
    'CL4': 'Used in Last Month',
    'CL5': 'Used in Last Week',
    'CL6': 'Used in Last Day',
    }
data['Alcohol'] = data['Alcohol'].replace(usage_col)
data['Amphet'] = data['Amphet'].replace(usage_col)
data['Amyl'] = data['Amyl'].replace(usage_col)
data['Benzos'] = data['Benzos'].replace(usage_col)
data['Caff'] = data['Caff'].replace(usage_col)
data['Cannabis'] = data['Cannabis'].replace(usage_col)
data['Choc'] = data['Choc'].replace(usage_col)
data['Coke'] = data['Coke'].replace(usage_col)
data['Crack'] = data['Crack'].replace(usage_col)
data['Ecstasy'] = data['Ecstasy'].replace(usage_col)
data['Heroin'] = data['Heroin'].replace(usage_col)
data['Ketamine'] = data['Ketamine'].replace(usage_col)
data['Legalh'] = data['Legalh'].replace(usage_col)
data['LSD'] = data['LSD'].replace(usage_col)
data['Meth'] = data['Meth'].replace(usage_col)
data['Mushrooms'] = data['Mushrooms'].replace(usage_col)
data['Nicotine'] = data['Nicotine'].replace(usage_col)
data['Semer'] = data['Semer'].replace(usage_col)
data['VSA'] = data['VSA'].replace(usage_col)

# %%
#rename column Choc to Chocolate
data = data.rename(columns = {'Choc':'Chocolate', 'Amyl':'Amyl Nitrite', 'Benzos':'Benzo Diazepine', 'Caff':'Caffeine', 'Coke':'Cocaine', 'Crack':'Crack Cocaine', 'Legalh':'Legal High', 'Meth':'Methadone', 'Mushrooms':'Magic Mushrooms', 'Semer':'Semeron', 'VSA':'Volatile Substance Abuse'})

# %%
drugs = ['Alcohol', 'Amphet', 'Amyl Nitrite', 'Benzo Diazepine', 'Caffeine', 'Cannabis', 'Chocolate', 'Cocaine', 'Crack Cocaine', 'Ecstasy', 'Heroin', 'Ketamine', 'Legal High', 'LSD', 'Methadone', 'Magic Mushrooms', 'Nicotine', 'Semeron', 'Volatile Substance Abuse']
drug_counts = pd.DataFrame()
drug_counts = drug_counts.rename_axis('Frequency')
for drug in drugs:
    counts = data[drug].value_counts()
    drug_counts[drug] = counts
drug_counts = drug_counts.fillna(0)

# %%
# Define the new order of rows
new_order = ['Used in Last Day'] + [index for index in drug_counts.index if index != 'Used in Last Day']
# Reindex the DataFrame
drug_counts = drug_counts.reindex(new_order)

# %%
from openpyxl import Workbook
from openpyxl.styles import PatternFill

wb = Workbook()

# Create a new worksheet and set its title
ws = wb.create_sheet("Frequence_usage")

# Add the DataFrame's header to the worksheet first
ws.append(['Frequency'] + drug_counts.columns.tolist())

# Then add the rest of the data
for index, row in drug_counts.iterrows():
    ws.append([index] + row.tolist())

for row in ws.iter_rows(min_row=1, max_row=drug_counts.shape[0]+1, min_col=1, max_col=1):
    for cell in row:
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type = "solid")

wb.save('drug_counts.xlsx')

# %%
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

wb = load_workbook('drug_counts.xlsx')

# Select the 'Frequence_usage' sheet
ws = wb['Frequence_usage']

chart = BarChart()
chart.type = "col"
chart.style = 12
chart.title = "Drug usage frequency per drug"
chart.y_axis.title = 'Frequency'
chart.x_axis.title = 'Drug'
chart.width = 50
chart.grouping = "stacked"  # Make the chart a stacked bar chart

# Assuming the drug data is in the rows and the frequency names are in the first column
donnees = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row, max_col=ws.max_column)
cats = Reference(ws, min_col=2, min_row=1, max_row=1, max_col=ws.max_column)

chart.add_data(donnees, titles_from_data=True, from_rows=True)
chart.set_categories(cats)
chart.shape = 4
ws.add_chart(chart, "A10")

wb.save('drug_counts.xlsx')

# %%
def calculate_percentages(data, drugs, frequency):
    education_drug_counts = pd.DataFrame()
    education_drug_counts = education_drug_counts.rename_axis('Frequency')
    for drug in drugs:
        total_frequency = data[data[drug] == frequency].shape[0]
        counts = data[data[drug] == frequency]['Education'].value_counts()
        education_drug_counts[drug] = (counts / total_frequency) * 100
    education_drug_counts = education_drug_counts.fillna(0)
    
    # Specify the order of the categories
    categories = ['Left School Before 16 years', 'Left School at 16 years', 'Left School at 17 years', 'Left School at 18 years',
                    'Some College,No Certificate Or Degree', 'Professional Certificate/ Diploma', 
                    'University Degree', 'Masters Degree', 'Doctorate Degree']
    education_drug_counts.index = pd.Categorical(education_drug_counts.index, categories=categories, ordered=True)
    
    # Sort by the categorical index
    education_drug_counts = education_drug_counts.sort_index()
    
    # Reverse the order of the rows
    education_drug_counts = education_drug_counts.iloc[::-1]
    
    return education_drug_counts

# Example usage:
frequency = 'Used in Last Day'  # Replace with any frequency

# %%
df_education_drug_count = calculate_percentages(data, drugs, frequency)

# %%
# Write the DataFrame to an Excel file
df_education_drug_count.to_excel('education_drug_counts.xlsx')

wb = load_workbook('education_drug_counts.xlsx')

# Select the first sheet in the workbook
ws = wb.active

chart = BarChart()
chart.type = "bar"  # Change this to "bar" for a horizontal bar chart
chart.style = 2
chart.title = "Drug usage frequency per drug"
chart.y_axis.title = 'Drug'  # Swap the x and y axis titles
chart.x_axis.title = 'Frequency'
chart.width = 50
chart.grouping = "stacked"  # Make the chart a stacked bar chart

# Assuming the drug data is in the columns and the frequency names are in the first row
donnees = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row, max_col=ws.max_column)
cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

chart.add_data(donnees, titles_from_data=True)
chart.set_categories(cats)
chart.shape = 4
ws.add_chart(chart, "A10")

wb.save('education_drug_counts.xlsx')

# %%
# Load existing workbook
wb = load_workbook('drug_counts.xlsx')

# Create a new sheet named 'data'
ws = wb.create_sheet('data')

# Add the DataFrame column headers as the first row in the sheet
for i, column in enumerate(data.columns):
    ws.cell(row=1, column=i + 1, value=column)

# Write data to the sheet, starting from the second row
for index, row in data.iterrows():
    for i, value in enumerate(row):
        # Write the value to the cell
        # Note: openpyxl uses 1-based indexing, so add 1 to the row and column numbers
        ws.cell(row=index + 2, column=i + 1, value=value)  # Start from the second row

# Save the changes
wb.save('drug_counts.xlsx')

# %%
from openpyxl import load_workbook
from openpyxl.chart import PieChart, Reference
import pandas as pd

# Load existing workbook
wb = load_workbook('drug_counts.xlsx')

# Get the 'data' sheet
data_ws = wb['data']

# Load the data into a pandas DataFrame, skipping the first row
data = pd.DataFrame(list(data_ws.values)[1:])

# Count the frequency of each unique value in column B
counts = data[1].value_counts()

# Create a new sheet for the counts
counts_ws = wb.create_sheet('counts')
for i, (index, value) in enumerate(counts.items()):
    counts_ws.cell(row=i + 2, column=1, value=index)
    counts_ws.cell(row=i + 2, column=2, value=value)

# Create a new sheet for the pie chart
chart_ws = wb.create_sheet('pie_chart')

# Create a new pie chart
chart = PieChart()
chart.title = "Age des répondants"
chart.style = 26
chart.width = 15
chart.height = 10

# Define the data for the chart
labels = Reference(counts_ws, min_col=1, min_row=2, max_row=len(counts) + 1)
data = Reference(counts_ws, min_col=2, min_row=2, max_row=len(counts) + 1)

# Add the data to the chart
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

# Place the chart on the new sheet
chart_ws.add_chart(chart, "E1")

# Save the changes
wb.save('drug_counts.xlsx')

# %%
from openpyxl import load_workbook
from openpyxl.chart import PieChart, Reference
import pandas as pd

# Load existing workbook
wb = load_workbook('drug_counts.xlsx')

# Get the 'data' sheet
data_ws = wb['data']

# Load the data into a pandas DataFrame
data = pd.DataFrame(data_ws.values)

# Set the column names
data.columns = data.iloc[0]
data = data[1:]

# Filter the data for each gender
male_data = data[data['Gender'] == 'Male']
female_data = data[data['Gender'] == 'Female']

# Count the frequency of each unique value in the 'Age' column for each gender
male_counts = male_data['Age'].value_counts()
female_counts = female_data['Age'].value_counts()

# Create a new sheet for the counts
counts_ws = wb.create_sheet('counts')

# Write the male counts to the new sheet
for i, (index, value) in enumerate(male_counts.items()):
    counts_ws.cell(row=i + 1, column=1, value=index)
    counts_ws.cell(row=i + 1, column=2, value=value)

# Write the female counts to the new sheet, offset by the number of male counts
for i, (index, value) in enumerate(female_counts.items()):
    counts_ws.cell(row=i + 2 + len(male_counts), column=1, value=index)
    counts_ws.cell(row=i + 2 + len(male_counts), column=2, value=value)

# Get the 'pie_chart' sheet
pie_chart_ws = wb['pie_chart']

# Create new pie charts
male_chart = PieChart()
male_chart.title = "Age distribution - Male"
male_chart.style = 2
male_chart.width = 15
male_chart.height = 10

female_chart = PieChart()
female_chart.title = "Age distribution - Female"
female_chart.style = 2
female_chart.width = 15
female_chart.height = 10

# Define the data for the charts
male_labels = Reference(counts_ws, min_col=1, min_row=1, max_row=len(male_counts))
male_data = Reference(counts_ws, min_col=2, min_row=1, max_row=len(male_counts))

female_labels = Reference(counts_ws, min_col=1, min_row=1 + len(male_counts), max_row=len(male_counts) + len(female_counts))
female_data = Reference(counts_ws, min_col=2, min_row=1 + len(male_counts), max_row=len(male_counts) + len(female_counts))

# Add the data to the charts
male_chart.add_data(male_data, titles_from_data=True)
male_chart.set_categories(male_labels)

female_chart.add_data(female_data, titles_from_data=True)
female_chart.set_categories(female_labels)

# Place the charts on the 'pie_chart' sheet at the specified locations
pie_chart_ws.add_chart(male_chart, "E22")
pie_chart_ws.add_chart(female_chart, "P22")

# Save the changes
wb.save('drug_counts.xlsx')


